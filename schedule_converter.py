#!/usr/bin/env python3
"""
教务系统课程表转换器：Word XML (.doc/.docx) 和 Excel (.xls/.xlsx) → schedule.json

用法:
    python schedule_converter.py 课程表.doc
    python schedule_converter.py 课程表.xls output.json
    python schedule_converter.py 课程表.doc --stdout

依赖:
    .doc/.docx  — 标准库（zipfile/re），无需安装
    .xls        — pip install xlrd
    .xlsx       — pip install openpyxl
"""

import re
import json
import sys
from pathlib import Path

WEEKDAY_CN = ["星期日", "星期一", "星期二", "星期三", "星期四", "星期五", "星期六"]
WEEKDAY_EN = ["Sunday", "Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]

# 节次时间表（可按需修改）
PERIODS_DEFINITION = [
    {"index": 1,  "time": "08:00~08:45"},
    {"index": 2,  "time": "08:55~09:40"},
    {"index": 3,  "time": "10:10~10:55"},
    {"index": 4,  "time": "11:05~11:50"},
    {"index": 5,  "time": "14:20~15:05"},
    {"index": 6,  "time": "15:15~16:00"},
    {"index": 7,  "time": "16:30~17:15"},
    {"index": 8,  "time": "17:25~18:10"},
    {"index": 9,  "time": "19:00~19:45"},
    {"index": 10, "time": "19:55~20:40"},
    {"index": 11, "time": "20:50~21:35"},
]


# ─── 通用工具 ────────────────────────────────────────────────────────────────

def make_result(semester: str, schedule: dict) -> dict:
    return {
        "semester":           semester,
        "periods_definition": PERIODS_DEFINITION,
        "schedule":           schedule,
    }


def sort_schedule(schedule: dict) -> dict:
    return {
        day: sorted(entries, key=lambda e: e["periods"][0] if e["periods"] else 0)
        for day, entries in schedule.items()
    }


# ─── Doc / Docx 解析 ─────────────────────────────────────────────────────────

def _read_xml(filepath: str) -> str:
    if Path(filepath).suffix.lower() == ".docx":
        import zipfile
        with zipfile.ZipFile(filepath) as z:
            return z.read("word/document.xml").decode("utf-8")
    with open(filepath, "r", encoding="utf-8") as f:
        return f.read()


def _parse_xml_cell(cxml: str) -> dict:
    body       = re.sub(r'<w:tcPr>.*?</w:tcPr>', '', cxml, flags=re.DOTALL)
    gs         = re.search(r'<w:gridSpan w:val="(\d+)"', cxml)
    span       = int(gs.group(1)) if gs else 1
    vm_restart = bool(re.search(r'<w:vMerge w:val="restart"', cxml))
    vm_any     = bool(re.search(r'<w:vMerge', cxml))
    vmerge     = "restart" if vm_restart else ("continue" if vm_any else None)
    texts      = re.findall(r'<w:t[^>]*>(.*?)</w:t>', body, re.DOTALL)
    return {"text": "".join(texts).strip(), "span": span, "vmerge": vmerge}


def _expand_xml_row(row_xml: str) -> list[tuple[int, dict]]:
    result, col = [], 0
    for cxml in re.findall(r'<w:tc>(.*?)</w:tc>', row_xml, re.DOTALL):
        cell = _parse_xml_cell(cxml)
        result.append((col, cell))
        col += cell["span"]
    return result


def _extract_xml_semester(xml: str) -> str:
    for para in re.findall(r'<w:p[ >].*?</w:p>', xml, re.DOTALL):
        texts = re.findall(r'<w:t[^>]*>(.*?)</w:t>', para, re.DOTALL)
        text  = "".join(texts).strip()
        if re.search(r'\d{4}.*学期', text):
            return text
    return ""


def convert_doc(filepath: str) -> dict:
    xml = _read_xml(filepath)

    table_match = re.search(r'<w:tbl>(.*?)</w:tbl>', xml, re.DOTALL)
    if not table_match:
        raise ValueError("未找到课程表格，请检查文件格式")

    rows_xml = re.findall(r'<w:tr[ >].*?</w:tr>', table_match.group(1), re.DOTALL)

    # 表头 → 物理列号到星期的映射
    col_to_day: dict[int, str] = {}
    for col, cell in _expand_xml_row(rows_xml[0]):
        if cell["text"] in WEEKDAY_CN:
            day = WEEKDAY_EN[WEEKDAY_CN.index(cell["text"])]
            for c in range(col, col + cell["span"]):
                col_to_day[c] = day

    schedule: dict[str, list] = {d: [] for d in WEEKDAY_EN}
    active:   dict[int, dict] = {}

    for row_xml in rows_xml[1:]:
        row    = _expand_xml_row(row_xml)
        period = None
        m      = re.search(r'第(\d+)节', row[0][1]["text"])
        if m:
            period = int(m.group(1))

        for col, cell in row:
            if col not in col_to_day:
                continue
            day, text, vmerge = col_to_day[col], cell["text"], cell["vmerge"]

            if vmerge == "restart" and text:
                entry = {"periods_start": period, "periods_end": period, "content": text}
                active[col] = entry
                schedule[day].append(entry)
            elif vmerge == "continue":
                if col in active:
                    active[col]["periods_end"] = period
            else:
                active.pop(col, None)
                if text:
                    entry = {"periods_start": period, "periods_end": period, "content": text}
                    schedule[day].append(entry)

    final: dict[str, list] = {}
    for day in WEEKDAY_EN:
        entries = []
        for e in schedule[day]:
            s, en   = e["periods_start"], e["periods_end"]
            periods = list(range(s, en + 1)) if s and en else ([s] if s else [])
            entries.append({"periods": periods, "content": e["content"]})
        final[day] = entries

    return make_result(_extract_xml_semester(xml), sort_schedule(final))


# ─── XLS / XLSX 解析 ─────────────────────────────────────────────────────────

def _parse_period_str(period_str: str) -> list[int]:
    """'1－2' / '1-2' / '9－10' → [1,2] / [9,10]"""
    s = period_str.strip().replace('－', '-').replace('—', '-')
    m = re.match(r'(\d+)-(\d+)', s)
    if m:
        return list(range(int(m.group(1)), int(m.group(2)) + 1))
    m2 = re.match(r'(\d+)', s)
    if m2:
        return [int(m2.group(1))]
    return []


def _parse_xls_cell(text: str) -> list[str]:
    """
    把一个单元格的文本拆成多条课程 content。
    格式：\n课名\n教师\n周次[周]\n教室\n（多门课以空行分隔）
    返回：['课名/教师/周次[周]/教室', ...]
    """
    text = text.strip()
    if not text:
        return []

    results = []
    # 按空行分割多门课
    for block in re.split(r'\n{2,}', text):
        lines = [l.strip() for l in block.split('\n') if l.strip()]
        if not lines:
            continue
        # 用斜杠拼接，保持和 doc 格式一致的风格
        content = "/".join(lines)
        results.append(content)
    return results


def _read_xls_sheet(filepath: str):
    """返回 (sheet对象, 行数, 列数)，自动选 xlrd / openpyxl"""
    suffix = Path(filepath).suffix.lower()
    if suffix == ".xls":
        import xlrd
        wb = xlrd.open_workbook(filepath)
        ws = wb.sheets()[0]

        class XlrdShim:
            def __init__(self, ws):
                self._ws = ws
                self.nrows = ws.nrows
                self.ncols = ws.ncols
            def cell(self, r, c):
                return self._ws.cell_value(r, c)

        return XlrdShim(ws)
    else:
        from openpyxl import load_workbook
        wb = load_workbook(filepath, read_only=True, data_only=True)
        ws = wb.active
        rows = list(ws.iter_rows(values_only=True))

        class OpenpyxlShim:
            def __init__(self, rows):
                self._rows = rows
                self.nrows = len(rows)
                self.ncols = max((len(r) for r in rows), default=0)
            def cell(self, r, c):
                v = self._rows[r][c] if c < len(self._rows[r]) else None
                return v if v is not None else ''

        return OpenpyxlShim(rows)


def _extract_xls_semester(sheet) -> str:
    """从前两行扫描学年学期字符串"""
    for r in range(min(3, sheet.nrows)):
        for c in range(sheet.ncols):
            val = str(sheet.cell(r, c)).strip()
            # "2025-2026-2" 格式 或 "学年学期" 关键字
            m = re.search(r'(\d{4}-\d{4}-\d)', val)
            if m:
                return m.group(1)
    return ""


def convert_xls(filepath: str) -> dict:
    sheet = _read_xls_sheet(filepath)

    # 扫描表头行（含"星期"关键字的行）
    header_row = -1
    col_to_day: dict[int, str] = {}

    for r in range(sheet.nrows):
        for c in range(sheet.ncols):
            val = str(sheet.cell(r, c)).strip()
            if val in WEEKDAY_CN:
                header_row = r
                break
        if header_row >= 0:
            break

    if header_row < 0:
        raise ValueError("未找到星期表头行，请检查文件格式")

    for c in range(sheet.ncols):
        val = str(sheet.cell(header_row, c)).strip()
        if val in WEEKDAY_CN:
            col_to_day[c] = WEEKDAY_EN[WEEKDAY_CN.index(val)]

    schedule: dict[str, list] = {d: [] for d in WEEKDAY_EN}

    for r in range(header_row + 1, sheet.nrows):
        period_str = str(sheet.cell(r, 0)).strip()
        periods    = _parse_period_str(period_str)
        if not periods:
            continue  # 备注行等，跳过

        for c, day in col_to_day.items():
            val = str(sheet.cell(r, c)).strip()
            for content in _parse_xls_cell(val):
                schedule[day].append({"periods": periods, "content": content})

    return make_result(_extract_xls_semester(sheet), sort_schedule(schedule))


# ─── 统一入口 ─────────────────────────────────────────────────────────────────

def convert(filepath: str) -> dict:
    suffix = Path(filepath).suffix.lower()
    if suffix in (".doc", ".docx"):
        return convert_doc(filepath)
    elif suffix in (".xls", ".xlsx"):
        return convert_xls(filepath)
    else:
        raise ValueError(f"不支持的文件格式：{suffix}（支持 .doc .docx .xls .xlsx）")


def main():
    if len(sys.argv) < 2:
        print(f"用法: python {Path(sys.argv[0]).name} 课程表.[doc|xls] [output.json | --stdout]")
        sys.exit(1)

    result   = convert(sys.argv[1])
    json_str = json.dumps(result, ensure_ascii=False, indent=2)
    stdout   = len(sys.argv) > 2 and sys.argv[2] == "--stdout"
    out_path = None if stdout else (sys.argv[2] if len(sys.argv) > 2 else "schedule.json")

    if stdout:
        print(json_str)
    else:
        Path(out_path).write_text(json_str, encoding="utf-8")
        total = sum(len(v) for v in result["schedule"].values())
        print(f"✅ 已输出到 {out_path}（学期: {result['semester']}，共 {total} 条课程）")


if __name__ == "__main__":
    main()
